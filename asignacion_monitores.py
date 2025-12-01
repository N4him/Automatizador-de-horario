import pandas as pd
import re
from datetime import datetime

"""
SISTEMA DE ASIGNACIÓN DE MONITORES
Asigna automáticamente monitores a horarios de salas según disponibilidad
"""

# ========================================================
# CONFIGURACIÓN - AJUSTA ESTOS VALORES
# ========================================================

CONFIG = {
    "monitores": {
        "archivo": "DISPONIBILIDAD HORARIA MONITORES DE SALAS 2025-II.xlsx",
        "hoja": "Hoja 1",
        "header_row": 4,        # Fila donde están los encabezados
        "data_start_row": 5,    # Fila donde empiezan los datos
        
        # Nombres de columnas (ajustar si son diferentes)
        "col_nombre": "Nombre completo",
        "col_min": None,        # Ej: "MIN" si existe la columna
        "col_max": None,        # Ej: "MAX" si existe la columna
        
        # Valores por defecto para MIN y MAX
        "horas_min_default": 8,
        "horas_max_default": 20
    },
    
    "espacios": {
        # Tu Excel que YA tiene formato lista
        "archivo": "Horario_Salas.xlsx",
        "hoja": 0,  # Primera hoja (o nombre de la hoja)
        
        # Nombres de columnas en tu Excel (ajustar si son diferentes)
        "col_sala": "SALA",
        "col_dia": "DIA",
        "col_hora_inicio": "HORA_INICIO",
        "col_hora_fin": "HORA_FIN",
        "col_curso": "CURSO"
    },
    
    "asignacion": {
        # Estrategias de asignación
        "balancear_carga": True,        # Distribuir equitativamente entre monitores
        "priorizar_minimo": True,       # Intentar que todos alcancen el mínimo primero
        
        # Restricciones adicionales
        "max_horas_seguidas": 4,        # Máximo de horas continuas (0 = sin límite)
        "descanso_minimo": 1,           # Horas de descanso entre turnos
        "permitir_sobrepasar_max": False  # Si True, puede asignar aunque exceda el máximo
    },
    
    "salida": {
        "archivo": "Asignacion_Monitores_Final.xlsx"
    }
}


# ========================================================
# FUNCIONES DE PARSEO
# ========================================================

def parse_time_str(time_str):
    """Convierte '7:00am' -> 7, '2:00pm' -> 14, '7' -> 7"""
    if pd.isna(time_str):
        return None
    
    s = str(time_str).strip().lower()
    
    # Formato AM/PM
    match = re.search(r'(\d{1,2})(?::(\d{2}))?\s*(am|pm)', s)
    if match:
        hour = int(match.group(1))
        meridiem = match.group(3)
        
        if meridiem == 'pm' and hour != 12:
            hour += 12
        elif meridiem == 'am' and hour == 12:
            hour = 0
        
        return hour
    
    # Solo número
    match = re.search(r'\d+', s)
    if match:
        return int(match.group(0))
    
    return None


def parse_range_cell(cell_value):
    """Convierte '7:00am-1:00pm' -> [(7, 13)]"""
    if pd.isna(cell_value):
        return []
    
    s = str(cell_value).strip().lower()
    
    # Casos especiales
    if s in ["libre", "disponible", "todo el día", "todo el dia"]:
        return [(7, 22)]
    
    if s in ["no disponible", "no", "n/a", "", "nan"]:
        return []
    
    # Buscar rangos
    ranges = []
    pattern = r'(\d{1,2}(?::\d{2})?\s*(?:am|pm)?)\s*-\s*(\d{1,2}(?::\d{2})?\s*(?:am|pm)?)'
    
    matches = re.findall(pattern, s)
    for match in matches:
        start = parse_time_str(match[0])
        end = parse_time_str(match[1])
        
        if start is not None and end is not None:
            ranges.append((start, end))
    
    return ranges


def normalizar_dia(dia):
    """Normaliza nombres de días"""
    if pd.isna(dia):
        return None
    
    d = str(dia).strip().lower()
    d = d.replace('á', 'a').replace('é', 'e').replace('í', 'i').replace('ó', 'o').replace('ú', 'u')
    d = re.sub(r'[^a-z]', '', d)
    
    mapeo = {
        'lun': 'lunes',
        'mar': 'martes',
        'mie': 'miercoles',
        'jue': 'jueves',
        'vie': 'viernes',
        'sab': 'sabado',
        'dom': 'domingo'
    }
    
    for abrev, completo in mapeo.items():
        if d.startswith(abrev):
            return completo
    
    return d if len(d) >= 3 else None


# ========================================================
# CARGA DE MONITORES
# ========================================================

def cargar_monitores():
    """Carga la disponibilidad de los monitores"""
    cfg = CONFIG["monitores"]
    
    print(f"\n📂 Cargando monitores: {cfg['archivo']}")
    
    # Leer archivo completo
    df_raw = pd.read_excel(cfg["archivo"], sheet_name=cfg["hoja"], header=None)
    
    # Extraer estructura de días y jornadas
    dias_row = df_raw.iloc[3]
    jornadas_row = df_raw.iloc[cfg["header_row"]]
    
    # Mapear columnas
    col_mapping = {}
    current_dia = None
    col_nombre_idx = None
    col_min_idx = None
    col_max_idx = None
    
    for idx, val in enumerate(jornadas_row):
        val_str = str(val).strip()
        
        # Columnas especiales
        if val_str == cfg["col_nombre"]:
            col_nombre_idx = idx
        elif cfg["col_min"] and val_str == cfg["col_min"]:
            col_min_idx = idx
        elif cfg["col_max"] and val_str == cfg["col_max"]:
            col_max_idx = idx
        
        # Días
        dia_val = dias_row[idx] if idx < len(dias_row) else None
        if pd.notna(dia_val) and str(dia_val).strip():
            current_dia = normalizar_dia(dia_val)
        
        # Jornadas
        val_lower = val_str.lower()
        if val_lower in ['mañana', 'manana', 'tarde', 'noche']:
            if current_dia:
                key = f"{current_dia}_{val_lower}"
                col_mapping[key] = idx
    
    if col_nombre_idx is None:
        print(f"❌ Error: No se encuentra la columna '{cfg['col_nombre']}'")
        return []
    
    print(f"✓ Estructura: {len(col_mapping)} bloques día/jornada")
    
    # Leer monitores
    monitores = []
    
    for row_idx in range(cfg["data_start_row"], len(df_raw)):
        row = df_raw.iloc[row_idx]
        
        nombre = row[col_nombre_idx]
        if pd.isna(nombre) or str(nombre).strip() == "":
            continue
        
        # MIN y MAX
        horas_min = cfg["horas_min_default"]
        horas_max = cfg["horas_max_default"]
        
        if col_min_idx is not None:
            try:
                horas_min = int(row[col_min_idx]) if pd.notna(row[col_min_idx]) else horas_min
            except:
                pass
        
        if col_max_idx is not None:
            try:
                horas_max = int(row[col_max_idx]) if pd.notna(row[col_max_idx]) else horas_max
            except:
                pass
        
        mon = {
            "id": row_idx - cfg["data_start_row"],
            "nombre": str(nombre).strip(),
            "min": horas_min,
            "max": horas_max,
            "horas": 0,
            "disp": {},
            "asignaciones": []
        }
        
        # Disponibilidad por día
        dias_unicos = set(k.split('_')[0] for k in col_mapping.keys())
        
        for dia in dias_unicos:
            mon["disp"][dia] = []
            
            for jornada in ['mañana', 'manana', 'tarde', 'noche']:
                key = f"{dia}_{jornada}"
                if key in col_mapping:
                    col_idx = col_mapping[key]
                    ranges = parse_range_cell(row[col_idx])
                    mon["disp"][dia].extend(ranges)
        
        monitores.append(mon)
    
    print(f"✅ {len(monitores)} monitores cargados")
    
    return monitores


# ========================================================
# CARGA DE ESPACIOS
# ========================================================

def cargar_espacios():
    """Carga espacios desde Excel formato lista"""
    cfg = CONFIG["espacios"]
    
    print(f"\n📂 Cargando espacios: {cfg['archivo']}")
    
    df = pd.read_excel(cfg["archivo"], sheet_name=cfg["hoja"])
    
    # Validar columnas
    columnas_req = [cfg["col_sala"], cfg["col_dia"], cfg["col_hora_inicio"], 
                    cfg["col_hora_fin"], cfg["col_curso"]]
    
    for col in columnas_req:
        if col not in df.columns:
            print(f"❌ Error: Columna '{col}' no encontrada")
            print(f"   Columnas disponibles: {list(df.columns)}")
            return pd.DataFrame()
    
    # Normalizar días
    df['DIA_NORM'] = df[cfg["col_dia"]].apply(normalizar_dia)
    
    # Calcular duración
    df['DURACION'] = df[cfg["col_hora_fin"]] - df[cfg["col_hora_inicio"]]
    
    print(f"✅ {len(df)} horarios cargados")
    print(f"   Salas: {df[cfg['col_sala']].nunique()}")
    print(f"   Total horas: {df['DURACION'].sum()}")
    
    return df


# ========================================================
# LÓGICA DE ASIGNACIÓN
# ========================================================

def esta_disponible(monitor, dia, hora_inicio, hora_fin):
    """Verifica disponibilidad del monitor"""
    if dia not in monitor["disp"]:
        return False
    
    for r_inicio, r_fin in monitor["disp"][dia]:
        if hora_inicio >= r_inicio and hora_fin <= r_fin:
            return True
    
    return False


def verificar_restricciones(monitor, dia, hora_inicio, hora_fin):
    """Verifica restricciones adicionales"""
    cfg = CONFIG["asignacion"]
    
    if not cfg.get("max_horas_seguidas"):
        return True
    
    # Revisar asignaciones del mismo día
    for asig in monitor["asignaciones"]:
        if asig["dia"] == dia:
            # Verificar si hay solapamiento o continuidad
            if (hora_inicio <= asig["fin"] and hora_fin >= asig["inicio"]):
                duracion_total = max(hora_fin, asig["fin"]) - min(hora_inicio, asig["inicio"])
                if duracion_total > cfg["max_horas_seguidas"]:
                    return False
    
    return True


def asignar_monitores(monitores, df_espacios):
    """Algoritmo principal de asignación"""
    cfg_asig = CONFIG["asignacion"]
    cfg_esp = CONFIG["espacios"]
    
    print(f"\n🔄 Iniciando asignación...")
    
    asignaciones = []
    sin_monitor = []
    
    espacios = df_espacios.to_dict('records')
    
    # Primera pasada: priorizar alcanzar mínimo
    if cfg_asig.get("priorizar_minimo"):
        print("   Fase 1: Alcanzando mínimos...")
        
        for espacio in espacios:
            dia = espacio['DIA_NORM']
            inicio = espacio[cfg_esp["col_hora_inicio"]]
            fin = espacio[cfg_esp["col_hora_fin"]]
            duracion = espacio['DURACION']
            
            # Buscar monitores bajo el mínimo
            candidatos = [
                m for m in monitores
                if m["horas"] < m["min"]
                and m["horas"] + duracion <= m["max"]
                and esta_disponible(m, dia, inicio, fin)
                and verificar_restricciones(m, dia, inicio, fin)
            ]
            
            if candidatos:
                # Asignar al que más le falta
                candidatos.sort(key=lambda x: x["min"] - x["horas"], reverse=True)
                elegido = candidatos[0]
                
                elegido["horas"] += duracion
                elegido["asignaciones"].append({
                    "dia": dia,
                    "inicio": inicio,
                    "fin": fin
                })
                
                asignaciones.append({
                    **espacio,
                    "MONITOR": elegido["nombre"],
                    "ESTADO": "✅"
                })
    
    # Segunda pasada: asignar restantes
    print("   Fase 2: Asignando restantes...")
    
    for espacio in espacios:
        # Verificar si ya fue asignado
        ya_asignado = any(
            a.get(cfg_esp["col_sala"]) == espacio[cfg_esp["col_sala"]] and
            a.get('DIA_NORM') == espacio['DIA_NORM'] and
            a.get(cfg_esp["col_hora_inicio"]) == espacio[cfg_esp["col_hora_inicio"]]
            for a in asignaciones
        )
        
        if ya_asignado:
            continue
        
        dia = espacio['DIA_NORM']
        inicio = espacio[cfg_esp["col_hora_inicio"]]
        fin = espacio[cfg_esp["col_hora_fin"]]
        duracion = espacio['DURACION']
        
        # Buscar candidatos
        candidatos = [
            m for m in monitores
            if m["horas"] + duracion <= m["max"]
            and esta_disponible(m, dia, inicio, fin)
            and verificar_restricciones(m, dia, inicio, fin)
        ]
        
        if not candidatos:
            sin_monitor.append(espacio)
            asignaciones.append({
                **espacio,
                "MONITOR": "SIN MONITOR",
                "ESTADO": "❌"
            })
            continue
        
        # Estrategia de balanceo
        if cfg_asig.get("balancear_carga"):
            candidatos.sort(key=lambda x: x["horas"])
        
        elegido = candidatos[0]
        elegido["horas"] += duracion
        elegido["asignaciones"].append({
            "dia": dia,
            "inicio": inicio,
            "fin": fin
        })
        
        asignaciones.append({
            **espacio,
            "MONITOR": elegido["nombre"],
            "ESTADO": "✅"
        })
    
    return asignaciones, sin_monitor


# ========================================================
# REPORTES
# ========================================================

def generar_reporte(monitores, asignaciones, sin_monitor):
    """Reporte en consola"""
    cfg_esp = CONFIG["espacios"]
    
    print("\n" + "="*70)
    print("📊 REPORTE DE ASIGNACIÓN")
    print("="*70)
    
    exitosos = len([a for a in asignaciones if a["ESTADO"] == "✅"])
    total = len(asignaciones)
    
    print(f"\n🎯 Resumen:")
    print(f"   Total horarios: {total}")
    print(f"   Asignados: {exitosos} ({exitosos*100/total:.1f}%)")
    print(f"   Sin monitor: {len(sin_monitor)} ({len(sin_monitor)*100/total:.1f}%)")
    
    # Por sala
    df = pd.DataFrame(asignaciones)
    print(f"\n🏢 Por Sala:")
    for sala in sorted(df[cfg_esp["col_sala"]].unique()):
        sala_df = df[df[cfg_esp["col_sala"]] == sala]
        asig = len(sala_df[sala_df['ESTADO'] == '✅'])
        print(f"   {sala:20} | {asig:3}/{len(sala_df):3} ({asig*100/len(sala_df):5.1f}%)")
    
    # Monitores
    print(f"\n👥 Monitores:")
    print("-" * 70)
    
    for m in sorted(monitores, key=lambda x: x["horas"], reverse=True):
        if m["horas"] > 0:
            pct = (m["horas"] / m["max"]) * 100
            barra = "█" * int(pct / 5)
            
            status = "✅"
            if m["horas"] < m["min"]:
                status = f"⚠️  <{m['min']}h"
            elif m["horas"] > m["max"]:
                status = f"❌ >{m['max']}h"
            
            print(f"{m['nombre'][:35]:35} | {m['horas']:2}h {barra:20} {status}")
    
    sin_carga = [m for m in monitores if m["horas"] == 0]
    if sin_carga:
        print(f"\n⚠️  Sin asignaciones ({len(sin_carga)}):")
        for m in sin_carga[:5]:
            print(f"   • {m['nombre']}")


def exportar_resultados(asignaciones, monitores):
    """Exporta a Excel en múltiples formatos"""
    cfg = CONFIG["salida"]
    cfg_esp = CONFIG["espacios"]
    
    df_asig = pd.DataFrame(asignaciones)
    
    # HOJA 1: Horarios visuales de TODAS las salas en una sola hoja
    # HOJA 2: Lista de asignaciones
    # HOJA 3: Monitores
    
    df_lista = df_asig.copy()
    
    df_mon = pd.DataFrame([{
        'Monitor': m['nombre'],
        'Horas': m['horas'],
        'Min': m['min'],
        'Max': m['max'],
        'Horarios': len(m['asignaciones']),
        'Estado': '✅' if m['min'] <= m['horas'] <= m['max'] else '⚠️'
    } for m in monitores]).sort_values('Horas', ascending=False)
    
    salas = sorted(df_asig[cfg_esp["col_sala"]].unique())
    
    with pd.ExcelWriter(cfg["archivo"], engine='openpyxl') as writer:
        # HOJA 1: Crear horario consolidado con todas las salas
        crear_horario_consolidado(writer, df_asig, salas, cfg_esp)
        
        # HOJA 2: Horario individual de cada monitor
        crear_horario_monitores(writer, monitores, df_asig, cfg_esp)
        
        # HOJA 3 y 4: Lista y resumen
        df_lista.to_excel(writer, sheet_name='Lista Asignaciones', index=False)
        df_mon.to_excel(writer, sheet_name='Resumen Monitores', index=False)
    
    print(f"\n✅ Exportado: {cfg['archivo']}")
    print(f"   📄 Hojas generadas:")
    print(f"      • Horarios Salas (todas las salas)")
    print(f"      • Horarios Monitores (detalle por monitor)")
    print(f"      • Lista Asignaciones")
    print(f"      • Resumen Monitores")


def crear_horario_monitores(writer, monitores, df_asig, cfg_esp):
    """Crea horario detallado de cada monitor en una sola hoja"""
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    # Crear hoja
    workbook = writer.book
    worksheet = workbook.create_sheet('Horarios Monitores', 1)
    
    # Estilos
    color_monitor = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    color_header = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    color_clase = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
    color_vacio = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    fuente_monitor = Font(bold=True, size=12, color="FFFFFF")
    fuente_header = Font(bold=True, size=10, color="000000")
    fuente_normal = Font(size=8, color="000000")
    
    alineacion_centro = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    borde = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    dias = ['Lunes', 'Martes', 'Miercoles', 'Jueves', 'Viernes', 'Sabado']
    
    # Determinar rango de horas
    hora_min = int(df_asig[cfg_esp["col_hora_inicio"]].min())
    hora_max = int(df_asig[cfg_esp["col_hora_fin"]].max())
    
    fila_actual = 1
    
    # Filtrar solo monitores con asignaciones
    monitores_activos = [m for m in monitores if m["horas"] > 0]
    monitores_activos.sort(key=lambda x: x["nombre"])
    
    for monitor in monitores_activos:
        # Título con nombre del monitor y total de horas
        worksheet.merge_cells(start_row=fila_actual, start_column=1, 
                             end_row=fila_actual, end_column=7)
        cell_titulo = worksheet.cell(row=fila_actual, column=1)
        cell_titulo.value = f"{monitor['nombre']} - {monitor['horas']} horas"
        cell_titulo.fill = color_monitor
        cell_titulo.font = fuente_monitor
        cell_titulo.alignment = alineacion_centro
        cell_titulo.border = borde
        fila_actual += 1
        
        # Encabezados de días
        worksheet.cell(row=fila_actual, column=1).value = "Hora"
        for idx, dia in enumerate(dias, start=2):
            cell = worksheet.cell(row=fila_actual, column=idx)
            cell.value = dia
            cell.fill = color_header
            cell.font = fuente_header
            cell.alignment = alineacion_centro
            cell.border = borde
        
        cell_hora_header = worksheet.cell(row=fila_actual, column=1)
        cell_hora_header.fill = color_header
        cell_hora_header.font = fuente_header
        cell_hora_header.alignment = alineacion_centro
        cell_hora_header.border = borde
        fila_actual += 1
        
        # Obtener asignaciones del monitor
        asignaciones_monitor = df_asig[df_asig['MONITOR'] == monitor['nombre']]
        
        # Crear matriz de horarios
        for hora in range(hora_min, hora_max):
            hora_str = f"{hora}-{hora+1}"
            
            # Columna de hora
            cell_hora = worksheet.cell(row=fila_actual, column=1)
            cell_hora.value = hora_str
            cell_hora.fill = color_header
            cell_hora.font = fuente_header
            cell_hora.alignment = alineacion_centro
            cell_hora.border = borde
            
            # Procesar cada día
            for idx_dia, dia in enumerate(dias, start=2):
                cell = worksheet.cell(row=fila_actual, column=idx_dia)
                cell.border = borde
                cell.alignment = alineacion_centro
                
                # Buscar asignaciones para este día/hora
                dia_norm = dia.lower()
                asig_celda = asignaciones_monitor[
                    (asignaciones_monitor['DIA_NORM'] == dia_norm) &
                    (asignaciones_monitor[cfg_esp["col_hora_inicio"]] <= hora) &
                    (asignaciones_monitor[cfg_esp["col_hora_fin"]] > hora)
                ]
                
                if len(asig_celda) > 0:
                    asig = asig_celda.iloc[0]
                    curso = asig[cfg_esp["col_curso"]]
                    sala = asig[cfg_esp["col_sala"]]
                    
                    cell.value = f"{sala}\n{curso}"
                    cell.fill = color_clase
                    cell.font = fuente_normal
                else:
                    cell.value = ""
                    cell.fill = color_vacio
            
            fila_actual += 1
        
        # Fila vacía entre monitores
        fila_actual += 2
    
    # Ajustar anchos
    worksheet.column_dimensions['A'].width = 10
    for col in range(2, 8):
        worksheet.column_dimensions[get_column_letter(col)].width = 22
    
    # Ajustar altura
    for row in range(1, fila_actual):
        worksheet.row_dimensions[row].height = 35


def crear_horario_consolidado(writer, df_asig, salas, cfg_esp):
    """Crea horario visual con todas las salas HORIZONTALMENTE (lado a lado)"""
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    workbook = writer.book
    worksheet = workbook.create_sheet('Horarios', 0)
    
    # Estilos
    color_naranja = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
    color_verde = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
    color_azul = PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid")
    color_vacio = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    color_sin_monitor = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
    color_header = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    color_titulo_sala = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    
    fuente_negra = Font(bold=False, size=8, color="000000")
    fuente_header = Font(bold=True, size=9, color="000000")
    fuente_titulo = Font(bold=True, size=11, color="FFFFFF")
    
    alineacion_centro = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    borde = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    dias = ['Lunes', 'Martes', 'Miercoles', 'Jueves', 'Viernes', 'Sabado']
    
    # Rango de horas
    hora_min = int(df_asig[cfg_esp["col_hora_inicio"]].min())
    hora_max = int(df_asig[cfg_esp["col_hora_fin"]].max())
    
    # ESTRUCTURA HORIZONTAL:
    # Fila 1: Títulos de salas (cada sala ocupa 7 columnas: 1 hora + 6 días)
    # Fila 2: Días de la semana (repetidos para cada sala)
    # Fila 3+: Horas + celdas de horario
    
    columna_actual = 1
    
    # Fila 1: Títulos de salas
    for sala in salas:
        # Título de sala (merge de 7 columnas)
        worksheet.merge_cells(
            start_row=1, 
            start_column=columna_actual, 
            end_row=1, 
            end_column=columna_actual + 6
        )
        cell_titulo = worksheet.cell(row=1, column=columna_actual)
        cell_titulo.value = sala
        cell_titulo.fill = color_titulo_sala
        cell_titulo.font = fuente_titulo
        cell_titulo.alignment = alineacion_centro
        cell_titulo.border = borde
        
        # Fila 2: Encabezados de días
        cell_hora_header = worksheet.cell(row=2, column=columna_actual)
        cell_hora_header.value = "Hora"
        cell_hora_header.fill = color_header
        cell_hora_header.font = fuente_header
        cell_hora_header.alignment = alineacion_centro
        cell_hora_header.border = borde
        
        for idx_dia, dia in enumerate(dias, start=1):
            cell = worksheet.cell(row=2, column=columna_actual + idx_dia)
            cell.value = dia
            cell.fill = color_header
            cell.font = fuente_header
            cell.alignment = alineacion_centro
            cell.border = borde
        
        columna_actual += 7  # Siguiente sala
    
    # Filas 3+: Horas y datos
    for fila_hora, hora in enumerate(range(hora_min, hora_max), start=3):
        hora_str = f"{hora}:00-{hora+1}:00"
        
        columna_actual = 1
        
        for sala in salas:
            df_sala = df_asig[df_asig[cfg_esp["col_sala"]] == sala]
            
            # Columna de hora
            cell_hora = worksheet.cell(row=fila_hora, column=columna_actual)
            cell_hora.value = hora_str
            cell_hora.fill = color_header
            cell_hora.font = fuente_header
            cell_hora.alignment = alineacion_centro
            cell_hora.border = borde
            
            # Procesar cada día
            for idx_dia, dia in enumerate(dias, start=1):
                cell = worksheet.cell(row=fila_hora, column=columna_actual + idx_dia)
                cell.border = borde
                cell.alignment = alineacion_centro
                
                dia_norm = dia.lower()
                asignaciones_celda = df_sala[
                    (df_sala['DIA_NORM'] == dia_norm) &
                    (df_sala[cfg_esp["col_hora_inicio"]] <= hora) &
                    (df_sala[cfg_esp["col_hora_fin"]] > hora)
                ]
                
                if len(asignaciones_celda) > 0:
                    asig = asignaciones_celda.iloc[0]
                    curso = asig[cfg_esp["col_curso"]]
                    monitor = asig['MONITOR']
                    
                    if monitor == "SIN MONITOR":
                        cell.value = f"{curso}\n❌ SIN MONITOR"
                        cell.fill = color_sin_monitor
                        cell.font = Font(size=7, color="FF0000", bold=True)
                    else:
                        nombre_corto = monitor.split()[0] if monitor else ""
                        cell.value = f"{curso}\n{nombre_corto}"
                        cell.font = fuente_negra
                        
                        hash_val = hash(monitor) % 3
                        if hash_val == 0:
                            cell.fill = color_naranja
                        elif hash_val == 1:
                            cell.fill = color_verde
                        else:
                            cell.fill = color_azul
                else:
                    cell.value = ""
                    cell.fill = color_vacio
            
            columna_actual += 7  # Siguiente sala
    
    # Ajustar anchos de columna
    for col_num in range(1, columna_actual):
        col_letter = get_column_letter(col_num)
        if (col_num - 1) % 7 == 0:  # Columnas de hora
            worksheet.column_dimensions[col_letter].width = 11
        else:  # Columnas de días
            worksheet.column_dimensions[col_letter].width = 18
    
    # Ajustar altura de filas
    worksheet.row_dimensions[1].height = 25  # Títulos
    worksheet.row_dimensions[2].height = 20  # Días
    for row in range(3, fila_hora + 1):
        worksheet.row_dimensions[row].height = 35


# ========================================================
# MAIN
# ========================================================

if __name__ == "__main__":
    print("="*70)
    print("🚀 SISTEMA DE ASIGNACIÓN DE MONITORES")
    print("="*70)
    
    try:
        monitores = cargar_monitores()
        df_espacios = cargar_espacios()
        
        if not monitores or len(df_espacios) == 0:
            print("\n❌ No se cargaron datos")
            exit(1)
        
        asignaciones, sin_monitor = asignar_monitores(monitores, df_espacios)
        generar_reporte(monitores, asignaciones, sin_monitor)
        exportar_resultados(asignaciones, monitores)
        
        print("\n✨ ¡Completado!")
        
    except Exception as e:
        print(f"\n❌ Error: {e}")
        import traceback
        traceback.print_exc()